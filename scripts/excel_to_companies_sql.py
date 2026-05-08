# -*- coding: utf-8 -*-
"""
엑셀(지사명·대표사업소, 협력회사명·낙찰예정자-상호 등, 사업자번호, 대표명, 주소) → companies INSERT SQL 생성.

사전 준비:
  pip install openpyxl

사용:
  python scripts/excel_to_companies_sql.py ./companies.xlsx -o companies_import.sql
  python scripts/excel_to_companies_sql.py ./companies2.xlsx -o companies2_import.sql --all-sheets

  (Windows에서 `> 파일.sql` 리다이렉트는 기본 인코딩이 CP949라 한글이 깨집니다.
   반드시 -o 옵션으로 UTF-8 저장하거나: $env:PYTHONIOENCODING='utf-8')

지사명: DB에는 "광산"처럼 저장, 엑셀은 "광산지사"인 경우 → 끝의 "지사"를 떼고 앞뒤 공백만 정리.
엑셀 표기가 DB `branches.name`과 다르면 BRANCH_NAME_ALIASES에 매핑 추가.

사업자번호: 하이픈·공백 제거 후 숫자만 저장(엑셀 숫자 셀도 처리).

한 회사(같은 사업자번호)가 여러 지사 행에 반복되는 것은 정상 → DB에도 branch_id만 다른 행으로 여러 개 들어감.
"""

from __future__ import annotations

import argparse
import io
import re
import sys
from pathlib import Path

# 엑셀/현장 표기 → DB `branches.name` (지정요일은 scripts/branch_allowed_weekdays_update.sql 참고)
BRANCH_NAME_ALIASES: dict[str, str] = {
    "광주전남본부직할": "본부 직할",
    "본부직할": "본부 직할",
}


def normalize_branch_name(raw: str) -> str:
    s = str(raw or "").strip()
    s = re.sub(r"\s+", " ", s)
    if s.endswith("지사"):
        s = s[: -len("지사")].strip()
    return BRANCH_NAME_ALIASES.get(s, s)


def normalize_business_registration_no(raw: object) -> str | None:
    """DB 기존 데이터처럼 숫자만 남김. 엑셀 숫자/문자·하이픈 혼용 처리."""
    if raw is None:
        return None
    if isinstance(raw, float) and not isinstance(raw, bool):
        if raw.is_integer():
            raw = int(raw)
        s = str(raw).strip()
    elif isinstance(raw, int) and not isinstance(raw, bool):
        s = str(raw)
    else:
        s = str(raw).strip()
    if not s:
        return None
    digits = re.sub(r"\D", "", s)
    return digits if digits else None


def sql_str(s: str | None) -> str:
    if s is None or str(s).strip() == "":
        return "null"
    t = str(s).strip()
    t = t.replace("'", "''")
    return "'" + t + "'"


def col_idx(header: tuple[object, ...], keywords: tuple[str, ...]) -> int:
    for i, cell in enumerate(header):
        if cell is None:
            continue
        h = str(cell).strip().replace(" ", "")
        for kw in keywords:
            if kw.replace(" ", "") in h or h in kw.replace(" ", ""):
                return i
    return -1


def process_sheet(
    ws,
    out,
    *,
    source_label: str,
) -> int:
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        print(f"-- 빈 시트 건너뜀: {source_label}", file=sys.stderr)
        return 0

    ci_branch = col_idx(header, ("대표사업소", "지사명", "사업소명", "지사"))
    ci_name = col_idx(header, ("낙찰예정자", "협력회사명", "회사명", "업체명"))
    ci_brn = col_idx(header, ("사업자번호", "사업자등록번호"))
    # '대표' 단독 키워드는 안 함 — '대표사업소' 헤더에 오탐
    ci_rep = col_idx(header, ("대표명", "대표자"))
    ci_addr = col_idx(header, ("주소",))

    missing = []
    if ci_branch < 0:
        missing.append("지사/대표사업소")
    if ci_name < 0:
        missing.append("업체명(낙찰예정자-상호 등)")
    if missing:
        print(f"헤더를 찾지 못했습니다 ({source_label}):", ", ".join(missing), file=sys.stderr)
        print("실제 헤더:", header, file=sys.stderr)
        raise ValueError("missing columns")

    out(f"-- sheet: {source_label}")
    n = 0
    for row in rows:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        branch_raw = row[ci_branch] if ci_branch < len(row) else None
        name = row[ci_name] if ci_name < len(row) else None
        if not branch_raw or not name:
            continue
        brn_raw = row[ci_brn] if ci_brn >= 0 and ci_brn < len(row) else None
        rep = row[ci_rep] if ci_rep >= 0 and ci_rep < len(row) else None
        addr = row[ci_addr] if ci_addr >= 0 and ci_addr < len(row) else None

        brn = normalize_business_registration_no(brn_raw)
        bnorm = normalize_branch_name(str(branch_raw))
        line = (
            "insert into public.companies (branch_id, name, business_registration_no, representative_name, address) "
            f"select b.id, {sql_str(name)}, {sql_str(brn)}, {sql_str(rep)}, {sql_str(addr)} "
            f"from public.branches b where b.name = {sql_str(bnorm)} "
            "on conflict (branch_id, name) do update set "
            "business_registration_no = excluded.business_registration_no, "
            "representative_name = excluded.representative_name, "
            "address = excluded.address, "
            "is_active = true;"
        )
        out(line)
        n += 1
    return n


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl이 필요합니다: pip install openpyxl", file=sys.stderr)
        return 1

    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path, help=".xlsx 파일 경로")
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="저장할 .sql 경로 (UTF-8). Windows에서는 리다이렉트 대신 이 옵션 사용",
    )
    ap.add_argument("--sheet", default=None, help="시트 이름 (기본: 첫 시트, --all-sheets 와 함께 쓰지 말 것)")
    ap.add_argument(
        "--all-sheets",
        action="store_true",
        help="통합 문서의 모든 시트를 순서대로 처리",
    )
    args = ap.parse_args()

    out_lines: list[str] = []

    def out(s: str = "") -> None:
        out_lines.append(s)

    if args.output is None:
        if hasattr(sys.stdout, "reconfigure"):
            try:
                sys.stdout.reconfigure(encoding="utf-8", errors="replace")
            except Exception:
                sys.stdout = io.TextIOWrapper(
                    sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True
                )
        out_wrapped = False
    else:
        out_wrapped = True

    if args.all_sheets and args.sheet:
        print("--all-sheets 와 --sheet 는 동시에 사용하지 마세요.", file=sys.stderr)
        return 1

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    total = 0
    try:
        if args.all_sheets:
            if out_wrapped:
                out(f"-- generated from {args.xlsx} (all sheets)")
                out("begin;")
            else:
                print(f"-- generated from {args.xlsx} (all sheets)")
                print("begin;")
            for name in wb.sheetnames:
                ws = wb[name]
                try:
                    total += process_sheet(ws, out if out_wrapped else (lambda s: print(s)), source_label=name)
                except ValueError:
                    return 1
            if out_wrapped:
                out("commit;")
                args.output.write_text("\n".join(out_lines) + "\n", encoding="utf-8")
            else:
                print("commit;")
        else:
            ws = wb[args.sheet] if args.sheet else wb.active
            label = args.sheet or (wb.sheetnames[0] if wb.sheetnames else "active")
            if out_wrapped:
                out(f"-- generated from {args.xlsx}")
                out("begin;")
            else:
                print(f"-- generated from {args.xlsx}")
                print("begin;")
            try:
                if out_wrapped:
                    total = process_sheet(ws, out, source_label=label)
                else:

                    def out_print(s: str = "") -> None:
                        print(s)

                    total = process_sheet(ws, out_print, source_label=label)
            except ValueError:
                return 1
            if out_wrapped:
                out("commit;")
                args.output.write_text("\n".join(out_lines) + "\n", encoding="utf-8")
            else:
                print("commit;")
    finally:
        wb.close()

    print("-- rows:", total, file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
